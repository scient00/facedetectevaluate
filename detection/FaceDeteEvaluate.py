#!/usr/bin/env python
# -*- coding: utf-8 -*-
from detection.FaceDetectInput import Rect
import cv2,os,string
import numpy as np
import logging,datetime,pandas,os
from copy import deepcopy
from openpyxl import *
import matplotlib.pyplot as plt

class DetectEvaluate(object):
    def __init__(self,classes = []):
        self.__deteListInfo = []                                                        #检测结果
        self.__labeListInfo = []                                                        #标注结果
        self.__imageDirPath = None                                                      #图片路径
        self.__classes = classes                                                        #测试检测类别
        self.__image = None                                                             #图像
        self.__iou = [iou/10.0 for iou in range(1,11,1)]                                #设置初始化阈值
        self.__midResult = self.__createDict()                                          #测试中间结果(二维字典)

    #构建统计指标的字典
    def __createDict(self):
        try:
            midResult = {}
            for iou in self.__iou:
                # dict构造多维字典，每一个列表共用内存地址
                # temp = dict(zip(self.__classes, [[0,0,0]] * len(self.__classes)))
                temp = {}
                for classesname in self.__classes:
                    # 字典列表依次存储[精度数组（precision）,召回率数组(recall),平均精度(ap)]
                    temp[classesname] = [None,None,0.0]
                midResult[iou] = temp
            return midResult
        except:
            return {}

    # 设置输入数据
    def set(self, deteListInfo, labeListInfo, imageDirPath):
        self.__deteListInfo = deteListInfo
        self.__labeListInfo = labeListInfo
        self.__imageDirPath = imageDirPath

    # 计算矩形重合率
    def __iouRect(self, rect1=Rect(), rect2=Rect()):
        try:
            endX = max(rect1.getX() + rect1.getWidth(), rect2.getX() + rect2.getWidth())
            startX = min(rect1.getX(), rect2.getX())
            width = rect1.getWidth() + rect2.getWidth() - (endX - startX)

            endY = max(rect1.getY() + rect1.getHeight(), rect2.getY() + rect2.getHeight())
            startY = min(rect1.getY(), rect2.getY())
            height = rect1.getHeight() + rect2.getHeight() - (endY - startY)

            if float(width) <= 0.0 or float(height) <= 0.0:
                ratio = 0.0
            else:
                Area = width * height
                Area1 = rect1.getWidth() * rect1.getHeight()
                Area2 = rect2.getWidth() * rect2.getHeight()
                ratio = Area * 1.0 / (Area1 + Area2 - Area)

            return ratio
        except:
            return 0.0

    # 保存图像
    def __saveImage(self, fpath):
        try:
            # python3.x 保存中文路径错误解决方法
            cv2.imencode('.jpg', self.__image)[1].tofile(fpath)
            # python3.x 保存英文路径
            # cv2.imwrite(fpath, self.__image)
        except:
            logging.error('The image is null!!!')

    # 分类标注的类别，转换为相应的元组(确保不可修改,空数据也需要列出),存储顺序[x,y,width,height,tab,confience]
    def __sortLabeClasses(self,classname):
        try:
            dictLabelcls = {}
            for k in range(len(self.__labeListInfo)):
                frame = self.__labeListInfo[k].getImageName()
                vlabelist = []
                judge = False
                for z in range(self.__labeListInfo[k].getSize()):
                    if classname == self.__labeListInfo[k].getDetectInfo(z).getTypeName():
                        label = list(self.__labeListInfo[k].getDetectInfo(z).getRect().toTuple())
                        #标志位设置为True
                        label.append(True)
                        label.append(self.__labeListInfo[k].getDetectInfo(z).getConfence())
                        vlabelist.append(label)
                        judge = True
                if judge == True:
                    dictLabelcls[frame] = vlabelist

            return dictLabelcls

        except:
            logging.error()
            return {}

    # 分类检测的类别，转换为相应的元组(确保不可修改，空数据剔除)
    def __sortDeteClasses(self,classname):
        try:
            frames = []
            rects = []
            confiences = []
            for k in range(len(self.__deteListInfo)):
                frame = self.__deteListInfo[k].getImageName()
                for z in range(self.__deteListInfo[k].getSize()):
                    if classname == self.__deteListInfo[k].getDetectInfo(z).getTypeName():
                        rect = self.__deteListInfo[k].getDetectInfo(z).getRect()
                        cofience = self.__deteListInfo[k].getDetectInfo(z).getConfence()
                        frames.append(frame)
                        rects.append(rect.toTuple())
                        confiences.append(cofience)
            return tuple(frames),tuple(rects),tuple(confiences)
        except:
            logging.error('__sortClasses error!!!')
            return (),(),()

    def __vocAveragePrecision(self, recall, precision, use_07_metric=True):
        try:
            if use_07_metric:
                ap = 0.
                # 2010年以前按recall等间隔取11个不同点处的精度值做平均(0., 0.1, 0.2, …, 0.9, 1.0)
                for t in np.arange(0., 1.1, 0.1):
                    if np.sum(recall >= t) == 0:
                        p = 0
                    else:
                        # 取最大值等价于2010以后先计算包络线的操作，保证precise非减
                        p = np.max(precision[recall >= t])
                    ap = ap + p / 11.
            else:
                # 2010年以后取所有不同的recall对应的点处的精度值做平均
                # first append sentinel values at the end
                mrec = np.concatenate(([0.], recall, [1.]))
                mpre = np.concatenate(([0.], precision, [0.]))

                # 计算包络线，从后往前取最大保证precise非减
                for i in range(mpre.size - 1, 0, -1):
                    mpre[i - 1] = np.maximum(mpre[i - 1], mpre[i])

                # 找出所有检测结果中recall不同的点
                i = np.where(mrec[1:] != mrec[:-1])[0]

                # and sum (\Delta recall) * prec
                # 用recall的间隔对精度作加权平均
                ap = np.sum((mrec[i + 1] - mrec[i]) * mpre[i + 1])
            return ap
        except:
            logging.error('__vocAveragePrecision error!!!')
            return 0.0

    #计算每个类别对应的AP，mAP是所有类别AP的平均值
    def __evaluate_series(self, classname,iouThresh,use_07_metric=True):
        try:
            assert 0 < iouThresh <= 1
            frames, rects, confiences = self.__sortDeteClasses(classname)
            dictclsLabel = self.__sortLabeClasses(classname)
            # 将该类别的检测结果按照置信度大小降序排列(order原索引)
            order = np.argsort(-np.array(confiences))
            dete_confidences = np.array(confiences)[order]

            # 对应的检测目标所属的图像名(帧号)
            dete_frame = np.array(frames)[order]
            dete_rect = np.array(rects)[order, :]

            # 用于标记每个检测结果是tp还是fp
            tp = np.zeros(len(dete_frame))
            fp = np.zeros(len(dete_frame))
            labeclsNum = 0
            for frame in dictclsLabel:
                labeclsNum += len(dictclsLabel[frame])

            for k, frame in enumerate(dete_frame):
                if frame not in dictclsLabel:
                    continue
                #存储顺序[x,y,width,height,标记位，置信度]
                iouValue = []
                rect1 =Rect(dete_rect[k][0],dete_rect[k][1],dete_rect[k][2],dete_rect[k][3])
                for z in range(len(dictclsLabel[frame])):
                    rect2 = Rect(dictclsLabel[frame][z][0],dictclsLabel[frame][z][1],dictclsLabel[frame][z][2],dictclsLabel[frame][z][3])
                    ratio = self.__iouRect(rect1,rect2)
                    iouValue.append(ratio)
                if len(iouValue) <= 0:
                    continue
                ovMaxIou = max(iouValue)
                index = iouValue.index(ovMaxIou)
                # 如果最大的重叠度大于一定的阈值
                if ovMaxIou >= iouThresh:
                    # 未被匹配过
                    if dictclsLabel[frame][index][4] == True:
                        #将匹配成功的标记位设置为False
                        dictclsLabel[frame][index][4] = False
                        tp[k] = 1
                    else:
                        # 若之前有置信度更高的检测结果匹配过这个标注结果，则此次检测结果为fp
                        fp[k] = 1
                else:
                    fp[k] = 1

            #按置信度取不同数量检测结果时的累计fp和tp
            cfp = np.cumsum(fp)
            ctp = np.cumsum(tp)
            # 精度为取的所有检测结果中tp的比例
            #precision = ctp / (ctp + cfp + 1e-7)
            precision = ctp / (ctp + cfp + np.finfo(np.float64).eps)
            # 召回率为占所有真实目标数量的比例
            recall = ctp / float(labeclsNum)
            # 计算recall-precise曲线下面积（严格来说并不是面积）
            ap = self.__vocAveragePrecision(precision, recall, use_07_metric=True)

            return precision,recall,ap
        except:
            logging.error('__evaluate_series error!!!')
            return -1.0,-1.0,-1.0

    # 近似平均精度
    def __ApproximatedAveragePrecision(self):
        try:

            for cls in self.__classes:
                for iou in self.__iou:
                    precision, recall, ap = self.__evaluate_series(cls,iou)
                    self.__midResult[iou][cls] = [precision,recall,ap]
        except:
            logging.error('ApproximatedAveragePrecision error!!!')

    # 评估
    def evaluate(self):
        try:
            self.__ApproximatedAveragePrecision()
        except:
            logging.error('evaluate error!!!')

    # 在图像上绘制显示信息(供调试调用)
    def showInfo(self,waitKey = 1,SaveImage = False,ShowImage = True,savepath = 'TestResult'):
        try:
            if self.__imageDirPath[-1] == '/' or self.__imageDirPath[-1] == '\\':
                dirname = self.__imageDirPath.strip().split(self.__imageDirPath[-1])[-2]
            else:
                dirname = self.__imageDirPath.split('/')[-1]

            savepath = savepath + os.path.sep + dirname
            if not os.path.exists(savepath):
                os.makedirs(savepath)

            for k in range(len(self.__deteListInfo)):
                deteInfo = self.__deteListInfo[k]
                labeInfo = self.__labeListInfo[k]
                if deteInfo.getImageName().find('.jpg') != -1 or deteInfo.getImageName().find(
                        '.png') != -1 or deteInfo.getImageName().find('.bmp') != -1:
                    imagePath = self.__imageDirPath +os.path.sep + deteInfo.getImageName()
                else:
                    imagePath = self.__imageDirPath + os.path.sep + deteInfo.getImageName() + '.jpg'
                try:
                    self.__image = cv2.imread(imagePath)
                    try:
                        if self.__image == None:
                            # 读取中文路径图像
                            self.__image = cv2.imdecode(np.fromfile(imagePath, dtype=np.uint8), 1)
                    except:
                        cv2.waitKey(waitKey)
                except:
                    self.__image = np.zeros([360, 640], np.uint8)

                font = cv2.FONT_HERSHEY_COMPLEX
                imageName = 'imageName: ' + deteInfo.getImageName()
                cv2.rectangle(self.__image, (5, 3), (350, 40), (255, 0, 255), 2)
                cv2.putText(self.__image, imageName, (9, 16), font, 0.5, (0, 255, 255), 1, cv2.LINE_AA)
                cv2.putText(self.__image, 'labele:', (9, 32), font, 0.5, (0, 0, 255), 1, cv2.LINE_AA)
                points = [[70, 20], [150, 20], [150, 33], [70, 33]]
                pts = np.array(points, np.int32)
                cv2.fillPoly(self.__image, [pts], (0, 0, 255))
                # cv2.line(self.__image, (80, 60), (200, 60), (0, 0, 255), 2)
                cv2.putText(self.__image, 'detect:', (154, 32), font, 0.5, (0, 255, 0), 1, cv2.LINE_AA)
                points = [[222, 20], [302, 20], [302, 33], [222, 33]]
                pts = np.array(points, np.int32)
                cv2.fillPoly(self.__image, [pts], (0, 255, 0))
                # cv2.line(self.__image, (80, 80), (200, 80), (0, 255, 0), 2)
                cv2.putText(self.__image, 'ClassesItem:', (9, 55), font, 0.5, (255, 0, 0), 1,cv2.LINE_AA)
                cv2.rectangle(self.__image,(9, 60), (60, 60 + 20*len(self.__classes)),(255,255,0),1)
                for k in range(len(self.__classes)):
                    cv2.putText(self.__image,str(self.__classes[k]),(10,75 + k*20),font,0.5,(153,50,204),1,cv2.LINE_AA )

                for k in range(deteInfo.getSize()):
                    x = deteInfo.getDetectInfo(k).getRect().getX()
                    y = deteInfo.getDetectInfo(k).getRect().getY()
                    width = deteInfo.getDetectInfo(k).getRect().getWidth() + x
                    height = deteInfo.getDetectInfo(k).getRect().getHeight() + y
                    cv2.rectangle(self.__image, (int(x), int(y)), (int(width), int(height)), (0, 255, 0), 1)
                    # cv2.putText(self.__image,self.__deteInfo.getDetectInfo(k).getTypeName(),(int(x + 5),int(y + 20)),font,0.5,(0,255,0),1,cv2.LINE_AA)

                for k in range(labeInfo.getSize()):
                    x = labeInfo.getDetectInfo(k).getRect().getX()
                    y = labeInfo.getDetectInfo(k).getRect().getY()
                    width = labeInfo.getDetectInfo(k).getRect().getWidth() + x
                    height = labeInfo.getDetectInfo(k).getRect().getHeight() + y
                    cv2.rectangle(self.__image, (int(x), int(y)), (int(width), int(height)), (0, 0, 255), 1)
                    # cv2.putText(self.__image, self.__labeInfo.getDetectInfo(k).getTypeName(), (int(x + 5), int(y + 40)), font, 0.5,(0, 0, 255), 1, cv2.LINE_AA)

                if ShowImage == True:
                    cv2.imshow('detectShow', self.__image)
                    cv2.waitKey(1)
                if SaveImage == True:
                    if deteInfo.getImageName().find('.jpg') != -1 or deteInfo.getImageName().find('.png') != -1 or deteInfo.getImageName().find('.bmp') != -1 :
                        fpath = savepath + os.path.sep + deteInfo.getImageName()
                    else:
                        fpath = savepath + os.path.sep + deteInfo.getImageName() + '.jpg'
                    self.__saveImage(fpath)

        except:
            logging.error('__showInfo error!!!')

    # 生成CSV报表
    def generateCsvReports(self, csvpath='.', ious = [e/10.0 for e in range(3,11,1)] ):
        try:
            if not os.path.exists(csvpath):
                os.makedirs(csvpath)
            with open(csvpath + os.path.sep + 'evaluateTestResult.csv', 'w') as cvsfile:
                cvsfile.write('Approximated Average Precision(近似平均精度),\n')
                for name in ['iou'] + self.__classes:
                    cvsfile.write(name + ',')
                cvsfile.write('\n')

                for iou in ious:
                    cvsfile.write(str(iou) + ',')
                    for cls in self.__classes:
                        ap = self.__midResult[iou][cls][2]
                        cvsfile.write(str(ap) + ',')
                    cvsfile.write('\n')

                cvsfile.write('\n\n')
                outInfo = u'Precision(精度):,正确被检测到数量/实际检测到的数量\n'
                outInfo += u'Recall(召回率):,正确被检测到数量/应该被检测到的数量\n'
                outInfo +=u'AveragePrecision(平均正确率):,Precision与Recall曲线积分(曲线下面积)\n'
                cvsfile.write(outInfo)
                cvsfile.write('\n')
        except:
            logging.error('generateCsvReports error!!!')

    #生成表报
    def generateHtmlReports(self,htmlpath = '.',ious = [e/10.0 for e in range(3,11,1)]):
        try:
            if not os.path.exists(htmlpath):
                # os.mkdir(pathTemp)
                os.makedirs(htmlpath)
            with open(htmlpath + os.path.sep + 'AveragePrecision.html','w' ) as filehtml:
                result = []
                resultDict = {}
                index = 0
                titles = ['iou'] + self.__classes
                result.append(ious)
                for cls in self.__classes:
                    temp = []
                    for iou in ious:
                        ap = self.__midResult[iou][cls][2]
                        temp.append(ap)
                    result.append(temp)
                    del temp

                for title in titles:
                    resultDict[title] = result[index]
                    index = index + 1

                df = pandas.DataFrame(resultDict)
                df = df[titles]
                html = df.to_html(index=False)
                filehtml.write(html)
        except:
            logging.error('generateHtmlReports error!!!')

    # 生成Excel报表
    def generateExcelReports(self, excelpath='.', ious=[e / 10.0 for e in range(3, 11, 1)]):
        try:

            dictLetter = list(string.ascii_uppercase)
            if not os.path.exists(excelpath):
                os.makedirs(excelpath)

            excelfile = excelpath + os.path.sep + 'evaluateTestResult.xlsx'

            wb = Workbook()
            sheet = wb.active
            sheet.title = '平均精度'

            row = 2 + len(ious) + 2
            cols = len(['iou'] + self.__classes)

            #合并单元格
            sheet.merge_cells('A1:%s1' % dictLetter[cols])
            sheet.cell(row=1, column=1, value ='Approximated Average Precision(近似平均精度)')

            for k in range(len(['iou'] + self.__classes)):
                sheet.cell(row = 2,column = k+1,value = (['iou'] + self.__classes)[k])

            for k in range(len(ious)):
                sheet.cell(row = k+3,column = 1,value = ious[k] )

                for z in range(len(self.__classes)):
                    ap = self.__midResult[ious[k]][self.__classes[z]][2]
                    sheet.cell(row = k+3,column = z+2,value = ap)

            outInfo = u'Precision(精度):正确被检测到数量/实际检测到的数量\n'
            outInfo += u'Recall(召回率):正确被检测到数量/应该被检测到的数量\n'
            outInfo += u'AveragePrecision(平均正确率):Precision与Recall曲线积分(曲线下面积)\n'
            sheet.merge_cells('A%d:%s%d' % (row,dictLetter[cols],row + 3))
            sheet.cell(row=row, column = 1, value=outInfo)
            wb.save(excelfile)

        except:
            logging.error('generateExcelReports error!!!')

    #生成precision_Recall曲线
    def generatePreRecCurve(self,pach = '.',ious=[e / 10.0 for e in range(3, 11, 1)]):
        try:
            fig, axes = plt.subplots(figsize=(3, 18), nrows=len(ious), ncols=len(self.__classes))
            if not isinstance(axes, np.ndarray):
                axes = np.array(axes)
            axes = axes.reshape(len(ious), len(self.__classes))

            for k, iou in enumerate(ious):
                for z, cls in enumerate(self.__classes):
                    precision = self.__midResult[iou][cls][0]
                    recall = self.__midResult[iou][cls][1]
                    axes[k, z].plot(recall, precision,color = 'blue')

            for ax, cls in zip(axes[0], self.__classes):
                ax.set_title(cls)

            for ax in axes[-1]:
                ax.set_xlabel('recall')

            for ax, iou in zip(axes[:, 0], ious):
                ax.set_ylabel('precison (IOU=%.1f)' % iou,fontsize = 10)

            plt.legend(loc='center left', bbox_to_anchor=(1.04, 0.5))
            plt.tight_layout()

            plt.savefig(pach + os.path.sep + 'precision_recall.jpg',bbox_inches='tight')
            #plt.show()
            plt.close()

        except:
            logging.error('generatePreRecCurve error!!!')